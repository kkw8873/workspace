# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6,3,5,8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 데이터 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for s in scores: # 기존 성적 데이터 넣기
    ws.append(s)

for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    ws.cell(row=idx, column=8).value ="=SUM(B{}:G{})".format(idx, idx)

    grade = None
    if sum_val >= 90:
        grade ="A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    if score[1] < 5:
        grade = "F"
    
    ws.cell(row=idx, column=9).value = grade 





# 내가한거 
# for quiz in ws.iter_rows(): # 퀴즈2 점수를 10으로 수정
#     if quiz[3].value != 10 and isinstance(quiz[3].value, int):
#         quiz[3].value = 10

# for i in range(1, 11):
#     ws["H1"] = "총합"
#     ws["H{}".format(i+1)] =f"=SUM(B{i+1}:G{i+1})"
#     ws["I1"] = "성적 정보"
#     if ws["H{}".format(i+1)].value >= 90:      
#         ws["I{}".format(i+1)].value = "A"
#     elif 90 > ws["H{}".format(i+1)].value >= 80:
#         ws["I{}".format(i+1)].value = "B"
#     elif 80 > ws["H{}".format(i+1)].value >= 70:
#         ws["I{}".format(i+1)].value = "C"
#     if ws["B{}".format(i+1)].value < 5:
#         ws["I{}".format(i+1)].value = "F"

wb.save("scores.xlsx")